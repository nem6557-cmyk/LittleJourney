"""Experimental-workbook loader (openpyxl) with the exact campaign schema.

Schema (per the experimental rig export):
  * header is on **row 5** (1-indexed); data begins on row 6.
  * columns are referenced 0-indexed:

    ===  ==========================  ====================
    col  quantity                    unit
    ===  ==========================  ====================
    14   Experiment Data Points      index (skip if NaN)
    23   Liquid Temperature (pool)   degC
    29   q''                          W/cm^2
    30   T_surf                       degC
    31   Pressure                     (rig units)
    32   T_sat                        degC
    33   subcool                      K
    34   dT_sat                       K
    35   htc                          W/m^2.K
    38   Thermal Resistance           K/W
    ===  ==========================  ====================

Rows whose data-point index is non-numeric are skipped (header padding,
blank separators, notes). ``T_surf`` in these workbooks is itself derived from
a single thermocouple via a 1-D conduction assumption -- we ingest it as-is and
flag that provenance in :class:`RunData`.

This loader is *real*: drop the proprietary ``.xlsx`` files into the workspace
and it works. When they are absent the validation harness uses the clearly
labelled synthetic generator instead.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Sequence

import numpy as np

# 0-indexed column map.
COLS = {
    "index": 14,
    "T_pool_C": 23,
    "q_Wcm2": 29,
    "T_surf_C": 30,
    "pressure": 31,
    "T_sat_C": 32,
    "subcool": 33,
    "dT_sat": 34,
    "htc": 35,
    "Rth": 38,
}
HEADER_ROW = 5     # 1-indexed
DATA_START_ROW = 6


@dataclass
class RunData:
    """One experimental (or synthetic) run: arrays aligned by data point."""

    name: str
    source: str                       # 'experiment' | 'SYNTHETIC'
    q_flux: np.ndarray                # W/m^2
    T_surf: np.ndarray                # K
    T_sat: np.ndarray                 # K
    T_pool: np.ndarray                # K
    pressure: Optional[np.ndarray] = None
    subcool: Optional[np.ndarray] = None
    dT_sat: Optional[np.ndarray] = None
    htc: Optional[np.ndarray] = None
    Rth: Optional[np.ndarray] = None
    # metadata
    design: str = ""
    coolant_in_C: Optional[float] = None
    fill_fraction: Optional[float] = None
    tsurf_note: str = "T_surf derived from a single thermocouple (1-D conduction)"

    def __len__(self) -> int:
        return len(self.q_flux)

    @property
    def n_points(self) -> int:
        return len(self.q_flux)


@dataclass
class WorkbookSpec:
    """Which sheets in a workbook map to which (design, coolant, fill)."""

    path: str
    design: str
    sheets: Dict[str, dict] = field(default_factory=dict)
    # sheets: {sheet_name: {"coolant_in_C": 20, "fill_fraction": 0.40}}


def _is_number(x) -> bool:
    if x is None:
        return False
    try:
        float(x)
        return True
    except (TypeError, ValueError):
        return False


def load_sheet(
    path: str,
    sheet: str,
    design: str = "",
    coolant_in_C: Optional[float] = None,
    fill_fraction: Optional[float] = None,
) -> RunData:
    """Load one worksheet into a :class:`RunData` using the fixed schema."""
    from openpyxl import load_workbook as _xl_load

    wb = _xl_load(path, data_only=True, read_only=True)
    ws = wb[sheet]
    cols: Dict[str, List[float]] = {k: [] for k in COLS}
    for row in ws.iter_rows(min_row=DATA_START_ROW):
        idx_cell = row[COLS["index"]].value if len(row) > COLS["index"] else None
        if not _is_number(idx_cell):
            continue
        for key, c in COLS.items():
            val = row[c].value if len(row) > c else None
            cols[key].append(float(val) if _is_number(val) else np.nan)
    wb.close()

    q = np.asarray(cols["q_Wcm2"]) * 1e4                  # W/cm^2 -> W/m^2
    T_surf = np.asarray(cols["T_surf_C"]) + 273.15
    T_sat = np.asarray(cols["T_sat_C"]) + 273.15
    T_pool = np.asarray(cols["T_pool_C"]) + 273.15
    return RunData(
        name=sheet, source="experiment", q_flux=q, T_surf=T_surf,
        T_sat=T_sat, T_pool=T_pool,
        pressure=np.asarray(cols["pressure"]),
        subcool=np.asarray(cols["subcool"]),
        dT_sat=np.asarray(cols["dT_sat"]),
        htc=np.asarray(cols["htc"]),
        Rth=np.asarray(cols["Rth"]),
        design=design, coolant_in_C=coolant_in_C, fill_fraction=fill_fraction,
    )


def load_workbook(spec: WorkbookSpec) -> List[RunData]:
    """Load every configured sheet in a workbook."""
    runs = []
    for sheet, meta in spec.sheets.items():
        runs.append(load_sheet(
            spec.path, sheet, design=spec.design,
            coolant_in_C=meta.get("coolant_in_C"),
            fill_fraction=meta.get("fill_fraction"),
        ))
    return runs


# Canonical workbook specs from the campaign description (paths relative to the
# workspace; present these files to run a real validation).
WORKBOOK_33 = WorkbookSpec(
    path="Plain_Chip_data_33_tube_condenser.xlsx",
    design="33-tube",
    sheets={
        "Plain chip 20C T2": {"coolant_in_C": 20, "fill_fraction": 0.40},
        "Plain chip 30C": {"coolant_in_C": 30, "fill_fraction": 0.40},
        "Plain chip 40C": {"coolant_in_C": 40, "fill_fraction": 0.40},
        "Plain chip 50C": {"coolant_in_C": 50, "fill_fraction": 0.40},
        "Plain chip 60%": {"coolant_in_C": 20, "fill_fraction": 0.60},
        "Plain chip 80%": {"coolant_in_C": 20, "fill_fraction": 0.80},
    },
)
WORKBOOK_42 = WorkbookSpec(
    path="Plain_Chip_data_older_condenser.xlsx",
    design="42-tube",
    sheets={
        "Plain Chip 20C old": {"coolant_in_C": 20, "fill_fraction": 0.40},
        "Plain Chip 30C old": {"coolant_in_C": 30, "fill_fraction": 0.40},
        "Plain Chip 40C old": {"coolant_in_C": 40, "fill_fraction": 0.40},
        "Plain Chip 50C Old": {"coolant_in_C": 50, "fill_fraction": 0.40},
    },
)
